"""导出工具: Excel / JSON / Markdown."""
import io
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


PUBLIC_FIELDS = [
    "id", "title", "difficulty", "domain", "subdomain",
    "content", "rubric", "reference_answer",
    "source_type", "source_detail",
    "author_name",
    "reviewer_name",
    "submitted_at", "reviewed_at",
]

FULL_FIELDS = PUBLIC_FIELDS + ["review_comment"]


def _project(q: dict, role: str) -> dict:
    fields = FULL_FIELDS if role == "reviewer" else PUBLIC_FIELDS
    return {k: q.get(k) for k in fields}


def to_json_bytes(questions: list, role: str) -> bytes:
    payload = [_project(q, role) for q in questions]
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def to_markdown_bytes(questions: list, role: str) -> bytes:
    lines = ["# CGT Agent benchmark 题库导出", ""]
    lines.append(f"导出时间: {datetime.now(timezone.utc).isoformat(timespec='seconds')}  ")
    lines.append(f"导出数量: {len(questions)}  ")
    lines.append(f"导出身份: {'审核员(全量)' if role == 'reviewer' else '出题人(公开字段)'}")
    lines.append("")
    for q in questions:
        qp = _project(q, role)
        lines.append(f"## #{qp['id']} {qp['title']}")
        lines.append("")
        lines.append(f"- 难度: **{qp['difficulty']}**")
        lines.append(f"- 领域: {qp['domain']}" + (f" / {qp['subdomain']}" if qp.get('subdomain') else ""))
        lines.append(f"- 来源: {qp['source_type']}" + (f" ({qp['source_detail']})" if qp.get('source_detail') else ""))
        lines.append(f"- 出题人: {qp['author_name']}")
        if qp.get("reviewer_name"):
            lines.append(f"- 审核人: {qp['reviewer_name']}")
        lines.append(f"- 提交时间: {qp['submitted_at']}")
        if qp.get("reviewed_at"):
            lines.append(f"- 审核时间: {qp['reviewed_at']}")
        lines.append("")
        lines.append("### 题目正文")
        lines.append("")
        lines.append(qp["content"])
        lines.append("")
        lines.append("### 采分点 (Rubric)")
        lines.append("")
        for i, item in enumerate(qp.get("rubric") or [], 1):
            lines.append(f"{i}. [{item.get('score')} 分] {item.get('desc')}")
        lines.append("")
        if qp.get("reference_answer"):
            lines.append("### 参考答案")
            lines.append("")
            lines.append(qp["reference_answer"])
            lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines).encode("utf-8")


def to_xlsx_bytes(questions: list, role: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = [
        "ID", "标题", "难度", "领域大类", "领域小类", "题目正文",
        "采分点", "参考答案",
        "题目来源", "来源详情",
        "出题人姓名",
    ]
    headers += [
        "审核人姓名",
        "提交时间", "审核时间",
    ]
    if role == "reviewer":
        headers.append("审核意见")

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [6, 30, 8, 22, 18, 50, 40, 40, 14, 22, 12, 12, 20, 20]
    if role == "reviewer":
        widths.append(40)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row_idx, q in enumerate(questions, 2):
        rubric_text = "\n".join(
            f"[{it.get('score')}分] {it.get('desc')}" for it in (q.get("rubric") or [])
        )
        row = [
            q.get("id"),
            q.get("title"),
            q.get("difficulty"),
            q.get("domain"),
            q.get("subdomain") or "",
            q.get("content"),
            rubric_text,
            q.get("reference_answer") or "",
            q.get("source_type"),
            q.get("source_detail") or "",
            q.get("author_name"),
            q.get("reviewer_name") or "",
            q.get("submitted_at") or "",
            q.get("reviewed_at") or "",
        ]
        if role == "reviewer":
            row.append(q.get("review_comment") or "")
        for col_idx, v in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
